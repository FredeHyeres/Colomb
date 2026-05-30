"""
Router Concours — module participant FCF
Trois modes de saisie des arrivées : manuelle, import CSV/Excel, Benzing Live!

Constantes FCF — à mettre à jour en début de saison si la règlementation évolue :
"""
import os
import csv
import io
import uuid
import threading
from datetime import datetime, date as DateType, timedelta, time as TimeType
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload

from database import get_db
from models.concours import Concours, ConcoursEngagement, StatutConcours, CategorieConcours, StatutEngagement
from models.pigeon import Pigeon
from models.sante import Sante, TypeEvenement
from models.sport import TrainingSession, PigeonTrainingResult, SessionType
from schemas.concours import (
    ConcoursCreate, ConcoursUpdate, ConcoursResponse,
    EngagementCreate, EngagementUpdate, EngagementResponse,
    ArriveeUpdate, ResultatUpdate, StatutUpdate,
    EligibiliteResult, ImportArriveeResult,
    ASPigeonResponse, StatsPigeonResponse,
)

# ── Constantes FCF (ajuster chaque saison si nécessaire) ──────────────────────
FCF_BAGUE_FCI_ANNEE_MIN = 2014      # bagues logo FCI obligatoires depuis 2014
FCF_AGE_MAX = 12                    # âge maximum en années
FCF_VACCINATION_MOIS = 12           # vaccination valable 12 mois

# ── Config Benzing (depuis .env) ───────────────────────────────────────────────
BENZING_FILE_PATH = os.getenv("BENZING_FILE_PATH", "")
BENZING_POLL_INTERVAL_SEC = int(os.getenv("BENZING_POLL_INTERVAL_SEC", "5"))

# ── State Benzing (par concours_id) ───────────────────────────────────────────
_benzing_watchers: dict = {}         # concours_id → {"thread": Thread, "active": bool, "nb_lus": int}

router = APIRouter(tags=["Concours"])


# ═══════════════════════════════════════════════════════════════════════════════
# UTILITAIRES
# ═══════════════════════════════════════════════════════════════════════════════

def _calculer_vitesse(concours: Concours, heure_arrivee_str: str, correction_sec: int) -> Optional[float]:
    """Calcule la vitesse en m/min à partir des heures de lâcher et d'arrivée."""
    if not concours.heure_lacher or not concours.distance_m or not heure_arrivee_str:
        return None
    try:
        base = datetime.combine(concours.date, TimeType.fromisoformat(concours.heure_lacher))
        arrivee = datetime.combine(concours.date, TimeType.fromisoformat(heure_arrivee_str))
        if arrivee < base:
            arrivee += timedelta(days=1)
        arrivee_corrigee = arrivee + timedelta(seconds=correction_sec)
        temps_min = (arrivee_corrigee - base).total_seconds() / 60
        if temps_min <= 0:
            return None
        return round(concours.distance_m / temps_min, 2)
    except Exception:
        return None


async def _verifier_eligibilite(pigeon: Pigeon, concours: Concours, db: AsyncSession) -> EligibiliteResult:
    """Vérifie les règles FCF et retourne le résultat d'éligibilité."""
    alertes = []
    annee_concours = concours.date.year

    # 1. Âge
    age = annee_concours - pigeon.annee_naissance
    if age > FCF_AGE_MAX:
        alertes.append(f"Âge dépassé : {age} ans (max {FCF_AGE_MAX} ans FCF)")

    # 2. Bague FCI
    if pigeon.annee_naissance < FCF_BAGUE_FCI_ANNEE_MIN:
        alertes.append(
            f"Bague non conforme : né en {pigeon.annee_naissance} "
            f"(logo FCI obligatoire depuis {FCF_BAGUE_FCI_ANNEE_MIN})"
        )

    # 3. Vaccination à jour
    date_limite_vaccin = concours.date - timedelta(days=FCF_VACCINATION_MOIS * 30)
    r = await db.execute(
        select(Sante)
        .where(Sante.pigeon_id == pigeon.id)
        .where(Sante.type == TypeEvenement.vaccination)
        .where(Sante.date >= date_limite_vaccin)
        .order_by(Sante.date.desc())
        .limit(1)
    )
    if not r.scalar_one_or_none():
        alertes.append(f"Vaccination absente ou expirée (aucune dans les {FCF_VACCINATION_MOIS} derniers mois)")

    # 4. Titre de propriété
    if not getattr(pigeon, "titre_propriete", True):
        alertes.append("Titre de propriété manquant")

    return EligibiliteResult(eligible=len(alertes) == 0, alertes=alertes)


async def _upsert_training_result(
    db: AsyncSession,
    session_id: int,
    pigeon_id: str,
    return_time: Optional[float],
    internal_rank: Optional[int],
    existing_result_id: Optional[int] = None,
) -> PigeonTrainingResult:
    """Crée ou met à jour le PigeonTrainingResult lié à l'engagement."""
    if existing_result_id:
        r = await db.execute(select(PigeonTrainingResult).where(PigeonTrainingResult.id == existing_result_id))
        tr = r.scalar_one_or_none()
        if tr:
            tr.return_time = return_time
            tr.internal_rank = internal_rank
            return tr

    tr = PigeonTrainingResult(
        session_id=session_id,
        pigeon_id=pigeon_id,
        return_time=return_time,
        internal_rank=internal_rank,
    )
    db.add(tr)
    await db.flush()
    return tr


# ═══════════════════════════════════════════════════════════════════════════════
# CONCOURS — CRUD
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/", response_model=List[ConcoursResponse])
async def list_concours(
    statut: Optional[str] = None,
    annee: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
):
    q = select(Concours).options(
        selectinload(Concours.engagements).selectinload(ConcoursEngagement.pigeon)
    ).order_by(Concours.date.desc())

    if statut:
        try:
            q = q.where(Concours.statut == StatutConcours(statut))
        except ValueError:
            pass
    if annee:
        q = q.where(func.extract("year", Concours.date) == annee)

    result = await db.execute(q)
    return result.scalars().all()


@router.post("/", response_model=ConcoursResponse, status_code=201)
async def create_concours(data: ConcoursCreate, db: AsyncSession = Depends(get_db)):
    # Créer la TrainingSession race liée
    session = TrainingSession(
        date=data.date,
        session_type=SessionType.race,
        distance_km=round(data.distance_m / 1000, 2) if data.distance_m else None,
        notes=data.nom,
    )
    db.add(session)
    await db.flush()

    data_dict = data.model_dump()
    data_dict["source"] = data_dict.get("source") or "manuel"
    concours = Concours(
        **data_dict,
        session_id=session.id,
    )
    db.add(concours)
    await db.commit()
    await db.refresh(concours)

    r = await db.execute(
        select(Concours)
        .options(selectinload(Concours.engagements).selectinload(ConcoursEngagement.pigeon))
        .where(Concours.id == concours.id)
    )
    return r.scalar_one()


@router.get("/stats/as-pigeon", response_model=List[ASPigeonResponse])
async def stats_as_pigeon(annee: Optional[int] = None, db: AsyncSession = Depends(get_db)):
    """Classement AS pigeon pour une saison."""
    q = (
        select(ConcoursEngagement)
        .options(
            selectinload(ConcoursEngagement.pigeon),
            selectinload(ConcoursEngagement.concours),
        )
    )
    if annee:
        q = q.join(Concours).where(func.extract("year", Concours.date) == annee)

    result = await db.execute(q)
    engagements = result.scalars().all()

    # Agréger par pigeon
    pigeons_map: dict = {}
    for eng in engagements:
        pid = eng.pigeon_id
        if pid not in pigeons_map:
            pigeons_map[pid] = {
                "pigeon_id": pid,
                "matricule": eng.pigeon.matricule if eng.pigeon else pid,
                "nb_concours": 0,
                "nb_classes": 0,
                "vitesses": [],
                "scores_as": [],
                "points_as": 0,
            }
        d = pigeons_map[pid]
        d["nb_concours"] += 1
        if eng.vitesse_m_min:
            d["vitesses"].append(eng.vitesse_m_min)
        if eng.statut == StatutEngagement.rentre_classe:
            d["nb_classes"] += 1
            d["points_as"] += 1
            if eng.classement_officiel and eng.nb_engages_categorie:
                d["scores_as"].append(eng.classement_officiel * 1000 / eng.nb_engages_categorie)

    rows = []
    for d in pigeons_map.values():
        vitesses = d["vitesses"]
        scores = d["scores_as"]
        nb = d["nb_concours"]
        rows.append(ASPigeonResponse(
            pigeon_id=d["pigeon_id"],
            matricule=d["matricule"],
            nb_concours=nb,
            nb_classes=d["nb_classes"],
            taux_retour=round(d["nb_classes"] / nb * 100, 1) if nb else 0.0,
            vitesse_moyenne=round(sum(vitesses) / len(vitesses), 2) if vitesses else None,
            meilleure_vitesse=max(vitesses) if vitesses else None,
            points_as=d["points_as"],
            score_as=round(sum(scores) / len(scores), 2) if scores else None,
        ))

    rows.sort(key=lambda x: (-x.points_as, x.score_as or 9999))
    return rows


@router.get("/stats/pigeon/{pigeon_id}", response_model=StatsPigeonResponse)
async def stats_pigeon(pigeon_id: str, db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(Pigeon).where(Pigeon.id == pigeon_id))
    pigeon = r.scalar_one_or_none()
    if not pigeon:
        raise HTTPException(status_code=404, detail="Pigeon non trouvé")

    q = (
        select(ConcoursEngagement)
        .options(selectinload(ConcoursEngagement.concours))
        .where(ConcoursEngagement.pigeon_id == pigeon_id)
        .join(Concours)
        .order_by(Concours.date.desc())
    )
    result = await db.execute(q)
    engs = result.scalars().all()

    nb_rentres = sum(1 for e in engs if e.statut in (StatutEngagement.rentre_classe, StatutEngagement.rentre_non_classe))
    nb_classes = sum(1 for e in engs if e.statut == StatutEngagement.rentre_classe)
    nb_non_rentres = len(engs) - nb_rentres
    vitesses = [e.vitesse_m_min for e in engs if e.vitesse_m_min]
    annees = sorted({e.concours.date.year for e in engs if e.concours}, reverse=True)

    return StatsPigeonResponse(
        pigeon_id=pigeon_id,
        matricule=pigeon.matricule,
        annees=annees,
        nb_concours=len(engs),
        nb_classes=nb_classes,
        nb_rentres=nb_rentres,
        nb_non_rentres=nb_non_rentres,
        taux_retour=round(nb_rentres / len(engs) * 100, 1) if engs else 0.0,
        vitesse_moyenne=round(sum(vitesses) / len(vitesses), 2) if vitesses else None,
        meilleure_vitesse=max(vitesses) if vitesses else None,
        engagements=engs,
    )


@router.get("/{concours_id}", response_model=ConcoursResponse)
async def get_concours(concours_id: str, db: AsyncSession = Depends(get_db)):
    r = await db.execute(
        select(Concours)
        .options(selectinload(Concours.engagements).selectinload(ConcoursEngagement.pigeon))
        .where(Concours.id == concours_id)
    )
    concours = r.scalar_one_or_none()
    if not concours:
        raise HTTPException(status_code=404, detail="Concours non trouvé")
    return concours


@router.put("/{concours_id}", response_model=ConcoursResponse)
async def update_concours(concours_id: str, data: ConcoursUpdate, db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(Concours).where(Concours.id == concours_id))
    concours = r.scalar_one_or_none()
    if not concours:
        raise HTTPException(status_code=404, detail="Concours non trouvé")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        if key == "statut":
            try:
                value = StatutConcours(value)
            except ValueError:
                raise HTTPException(status_code=422, detail=f"Statut invalide : {value}")
        setattr(concours, key, value)

    # Sync TrainingSession si date ou distance changés
    if concours.session_id and ("date" in update_data or "distance_m" in update_data or "nom" in update_data):
        sr = await db.execute(select(TrainingSession).where(TrainingSession.id == concours.session_id))
        session = sr.scalar_one_or_none()
        if session:
            if "date" in update_data:
                session.date = update_data["date"]
            if "distance_m" in update_data and update_data["distance_m"]:
                session.distance_km = round(update_data["distance_m"] / 1000, 2)
            if "nom" in update_data:
                session.notes = update_data["nom"]

    await db.commit()
    r2 = await db.execute(
        select(Concours)
        .options(selectinload(Concours.engagements).selectinload(ConcoursEngagement.pigeon))
        .where(Concours.id == concours_id)
    )
    return r2.scalar_one()


@router.patch("/{concours_id}/statut", response_model=ConcoursResponse)
async def patch_statut_concours(concours_id: str, data: StatutUpdate, db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(Concours).where(Concours.id == concours_id))
    concours = r.scalar_one_or_none()
    if not concours:
        raise HTTPException(status_code=404, detail="Concours non trouvé")
    try:
        concours.statut = StatutConcours(data.statut)
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Statut invalide : {data.statut}")
    await db.commit()
    r2 = await db.execute(
        select(Concours)
        .options(selectinload(Concours.engagements).selectinload(ConcoursEngagement.pigeon))
        .where(Concours.id == concours_id)
    )
    return r2.scalar_one()


@router.delete("/{concours_id}", status_code=204)
async def delete_concours(concours_id: str, db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(Concours).where(Concours.id == concours_id))
    concours = r.scalar_one_or_none()
    if not concours:
        raise HTTPException(status_code=404, detail="Concours non trouvé")

    session_id = concours.session_id
    await db.delete(concours)

    # Supprimer la TrainingSession liée si elle existe
    if session_id:
        sr = await db.execute(select(TrainingSession).where(TrainingSession.id == session_id))
        session = sr.scalar_one_or_none()
        if session:
            await db.delete(session)

    await db.commit()


# ═══════════════════════════════════════════════════════════════════════════════
# ENGAGEMENTS
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/{concours_id}/engagements", response_model=List[EngagementResponse])
async def list_engagements(concours_id: str, db: AsyncSession = Depends(get_db)):
    r = await db.execute(
        select(ConcoursEngagement)
        .options(selectinload(ConcoursEngagement.pigeon))
        .where(ConcoursEngagement.concours_id == concours_id)
    )
    return r.scalars().all()


@router.get("/{concours_id}/eligibilite/{pigeon_id}", response_model=EligibiliteResult)
async def check_eligibilite(concours_id: str, pigeon_id: str, db: AsyncSession = Depends(get_db)):
    r_c = await db.execute(select(Concours).where(Concours.id == concours_id))
    concours = r_c.scalar_one_or_none()
    if not concours:
        raise HTTPException(status_code=404, detail="Concours non trouvé")

    r_p = await db.execute(select(Pigeon).where(Pigeon.id == pigeon_id))
    pigeon = r_p.scalar_one_or_none()
    if not pigeon:
        raise HTTPException(status_code=404, detail="Pigeon non trouvé")

    return await _verifier_eligibilite(pigeon, concours, db)


@router.post("/{concours_id}/engagements", response_model=EngagementResponse, status_code=201)
async def add_engagement(concours_id: str, data: EngagementCreate, db: AsyncSession = Depends(get_db)):
    r_c = await db.execute(select(Concours).where(Concours.id == concours_id))
    concours = r_c.scalar_one_or_none()
    if not concours:
        raise HTTPException(status_code=404, detail="Concours non trouvé")

    r_p = await db.execute(select(Pigeon).where(Pigeon.id == data.pigeon_id))
    pigeon = r_p.scalar_one_or_none()
    if not pigeon:
        raise HTTPException(status_code=404, detail="Pigeon non trouvé")

    # Vérification doublon
    r_dup = await db.execute(
        select(ConcoursEngagement)
        .where(ConcoursEngagement.concours_id == concours_id)
        .where(ConcoursEngagement.pigeon_id == data.pigeon_id)
    )
    if r_dup.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Ce pigeon est déjà engagé dans ce concours")

    # Éligibilité
    elig = await _verifier_eligibilite(pigeon, concours, db)

    try:
        cat = CategorieConcours(data.categorie)
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Catégorie invalide : {data.categorie}")

    # Créer un PigeonTrainingResult vide dans la session liée
    tr_id = None
    if concours.session_id:
        tr = await _upsert_training_result(db, concours.session_id, data.pigeon_id, None, None)
        tr_id = tr.id

    engagement = ConcoursEngagement(
        concours_id=concours_id,
        pigeon_id=data.pigeon_id,
        categorie=cat,
        mise=data.mise,
        notes=data.notes,
        eligible=elig.eligible,
        raison_ineligibilite="\n".join(elig.alertes) if elig.alertes else None,
        training_result_id=tr_id,
    )
    db.add(engagement)
    await db.commit()
    await db.refresh(engagement)

    r = await db.execute(
        select(ConcoursEngagement)
        .options(selectinload(ConcoursEngagement.pigeon))
        .where(ConcoursEngagement.id == engagement.id)
    )
    return r.scalar_one()


@router.delete("/{concours_id}/engagements/{eng_id}", status_code=204)
async def remove_engagement(concours_id: str, eng_id: str, db: AsyncSession = Depends(get_db)):
    r = await db.execute(
        select(ConcoursEngagement)
        .where(ConcoursEngagement.id == eng_id)
        .where(ConcoursEngagement.concours_id == concours_id)
    )
    eng = r.scalar_one_or_none()
    if not eng:
        raise HTTPException(status_code=404, detail="Engagement non trouvé")

    # Supprimer le PigeonTrainingResult lié si vide (pas d'arrivée enregistrée)
    if eng.training_result_id and not eng.heure_arrivee:
        tr_r = await db.execute(select(PigeonTrainingResult).where(PigeonTrainingResult.id == eng.training_result_id))
        tr = tr_r.scalar_one_or_none()
        if tr:
            await db.delete(tr)

    await db.delete(eng)
    await db.commit()


# ── MODE 1 : Saisie manuelle d'arrivée ─────────────────────────────────────────

@router.patch("/{concours_id}/engagements/{eng_id}/arrivee", response_model=EngagementResponse)
async def saisir_arrivee(
    concours_id: str, eng_id: str, data: ArriveeUpdate, db: AsyncSession = Depends(get_db)
):
    r = await db.execute(
        select(ConcoursEngagement)
        .options(selectinload(ConcoursEngagement.pigeon))
        .where(ConcoursEngagement.id == eng_id)
        .where(ConcoursEngagement.concours_id == concours_id)
    )
    eng = r.scalar_one_or_none()
    if not eng:
        raise HTTPException(status_code=404, detail="Engagement non trouvé")

    rc = await db.execute(select(Concours).where(Concours.id == concours_id))
    concours = rc.scalar_one()

    vitesse = _calculer_vitesse(concours, data.heure_arrivee, data.correction_horloge_sec)
    eng.heure_arrivee = data.heure_arrivee
    eng.correction_horloge_sec = data.correction_horloge_sec
    eng.vitesse_m_min = vitesse
    eng.source_arrivee = "manuelle"

    # Sync PigeonTrainingResult
    if concours.session_id:
        tr = await _upsert_training_result(
            db, concours.session_id, eng.pigeon_id,
            return_time=(vitesse and round(concours.distance_m / vitesse, 2) if concours.distance_m and vitesse else None),
            internal_rank=eng.classement_officiel,
            existing_result_id=eng.training_result_id,
        )
        eng.training_result_id = tr.id

    await db.commit()
    r2 = await db.execute(
        select(ConcoursEngagement)
        .options(selectinload(ConcoursEngagement.pigeon))
        .where(ConcoursEngagement.id == eng_id)
    )
    return r2.scalar_one()


# ── Résultat officiel ──────────────────────────────────────────────────────────

@router.patch("/{concours_id}/engagements/{eng_id}/resultat", response_model=EngagementResponse)
async def saisir_resultat(
    concours_id: str, eng_id: str, data: ResultatUpdate, db: AsyncSession = Depends(get_db)
):
    r = await db.execute(
        select(ConcoursEngagement)
        .options(selectinload(ConcoursEngagement.pigeon))
        .where(ConcoursEngagement.id == eng_id)
        .where(ConcoursEngagement.concours_id == concours_id)
    )
    eng = r.scalar_one_or_none()
    if not eng:
        raise HTTPException(status_code=404, detail="Engagement non trouvé")

    if data.classement_officiel is not None:
        eng.classement_officiel = data.classement_officiel
    if data.nb_engages_categorie is not None:
        eng.nb_engages_categorie = data.nb_engages_categorie
    if data.statut:
        try:
            eng.statut = StatutEngagement(data.statut)
        except ValueError:
            raise HTTPException(status_code=422, detail=f"Statut invalide : {data.statut}")

    # Sync internal_rank dans PigeonTrainingResult
    if eng.training_result_id and data.classement_officiel is not None:
        tr_r = await db.execute(select(PigeonTrainingResult).where(PigeonTrainingResult.id == eng.training_result_id))
        tr = tr_r.scalar_one_or_none()
        if tr:
            tr.internal_rank = data.classement_officiel

    await db.commit()
    r2 = await db.execute(
        select(ConcoursEngagement)
        .options(selectinload(ConcoursEngagement.pigeon))
        .where(ConcoursEngagement.id == eng_id)
    )
    return r2.scalar_one()


@router.patch("/{concours_id}/engagements/{eng_id}/statut", response_model=EngagementResponse)
async def patch_statut_engagement(
    concours_id: str, eng_id: str, data: StatutUpdate, db: AsyncSession = Depends(get_db)
):
    r = await db.execute(
        select(ConcoursEngagement)
        .options(selectinload(ConcoursEngagement.pigeon))
        .where(ConcoursEngagement.id == eng_id)
        .where(ConcoursEngagement.concours_id == concours_id)
    )
    eng = r.scalar_one_or_none()
    if not eng:
        raise HTTPException(status_code=404, detail="Engagement non trouvé")
    try:
        eng.statut = StatutEngagement(data.statut)
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Statut invalide : {data.statut}")
    await db.commit()
    r2 = await db.execute(
        select(ConcoursEngagement)
        .options(selectinload(ConcoursEngagement.pigeon))
        .where(ConcoursEngagement.id == eng_id)
    )
    return r2.scalar_one()


# ═══════════════════════════════════════════════════════════════════════════════
# MODE 2 : IMPORT CSV / EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

@router.post("/{concours_id}/import-arrivees", response_model=ImportArriveeResult)
async def import_arrivees(
    concours_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Import des arrivées depuis un fichier CSV ou Excel.
    Format CSV : matricule;heure_arrivee;correction_sec
    Format Excel : mêmes colonnes, premier onglet.
    """
    rc = await db.execute(select(Concours).where(Concours.id == concours_id))
    concours = rc.scalar_one_or_none()
    if not concours:
        raise HTTPException(status_code=404, detail="Concours non trouvé")

    content = await file.read()
    filename = (file.filename or "").lower()
    rows = []

    if filename.endswith(".xlsx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: str(v or "").strip() for i, v in enumerate(row) if i < len(headers)})
        except ImportError:
            raise HTTPException(status_code=400, detail="openpyxl non disponible pour les fichiers .xlsx")
    else:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        reader = csv.DictReader(io.StringIO(text), delimiter=";")
        rows = list(reader)

    traites, non_trouves, erreurs = 0, [], []

    # Indexer les engagements par matricule
    eng_result = await db.execute(
        select(ConcoursEngagement)
        .options(selectinload(ConcoursEngagement.pigeon))
        .where(ConcoursEngagement.concours_id == concours_id)
    )
    engs_by_matricule = {
        e.pigeon.matricule: e for e in eng_result.scalars().all() if e.pigeon
    }

    for i, row in enumerate(rows, start=2):
        matricule = (row.get("matricule") or "").strip()
        heure = (row.get("heure_arrivee") or "").strip()
        correction = int((row.get("correction_sec") or "0").strip() or "0")

        if not matricule or not heure:
            erreurs.append(f"Ligne {i} : matricule ou heure_arrivee manquant")
            continue

        eng = engs_by_matricule.get(matricule)
        if not eng:
            non_trouves.append(matricule)
            continue

        vitesse = _calculer_vitesse(concours, heure, correction)
        eng.heure_arrivee = heure
        eng.correction_horloge_sec = correction
        eng.vitesse_m_min = vitesse
        eng.source_arrivee = "csv"

        if concours.session_id:
            tr = await _upsert_training_result(
                db, concours.session_id, eng.pigeon_id,
                return_time=(round(concours.distance_m / vitesse, 2) if concours.distance_m and vitesse else None),
                internal_rank=eng.classement_officiel,
                existing_result_id=eng.training_result_id,
            )
            eng.training_result_id = tr.id

        traites += 1

    await db.commit()
    return ImportArriveeResult(traites=traites, non_trouves=non_trouves, erreurs=erreurs)


# ═══════════════════════════════════════════════════════════════════════════════
# MODE 3 : BENZING LIVE!
# ═══════════════════════════════════════════════════════════════════════════════

def _benzing_watch_loop(concours_id: str, file_path: str, state: dict):
    """Thread de polling du fichier log Benzing."""
    last_pos = 0
    while state["active"]:
        try:
            if os.path.exists(file_path):
                with open(file_path, "r", encoding="latin-1") as f:
                    f.seek(last_pos)
                    new_lines = f.readlines()
                    last_pos = f.tell()
                if new_lines:
                    state["pending"].extend([l.strip() for l in new_lines if l.strip()])
        except Exception as e:
            state["last_error"] = str(e)
        import time
        time.sleep(BENZING_POLL_INTERVAL_SEC)


@router.post("/{concours_id}/benzing/start-watch")
async def benzing_start(concours_id: str, db: AsyncSession = Depends(get_db)):
    if not BENZING_FILE_PATH:
        return {"available": False, "message": "BENZING_FILE_PATH non configuré dans .env"}

    rc = await db.execute(select(Concours).where(Concours.id == concours_id))
    if not rc.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Concours non trouvé")

    if concours_id in _benzing_watchers and _benzing_watchers[concours_id]["active"]:
        return {"available": True, "active": True, "message": "Watch déjà actif"}

    state = {"active": True, "nb_lus": 0, "pending": [], "last_error": None}
    t = threading.Thread(
        target=_benzing_watch_loop,
        args=(concours_id, BENZING_FILE_PATH, state),
        daemon=True,
    )
    t.start()
    _benzing_watchers[concours_id] = {"thread": t, "state": state}
    return {"available": True, "active": True, "message": f"Watch démarré sur {BENZING_FILE_PATH}"}


@router.get("/{concours_id}/benzing/status")
async def benzing_status(concours_id: str):
    if not BENZING_FILE_PATH:
        return {"available": False, "active": False, "message": "BENZING_FILE_PATH non configuré"}

    watcher = _benzing_watchers.get(concours_id)
    if not watcher:
        return {"available": True, "active": False, "nb_lus": 0}

    state = watcher["state"]
    return {
        "available": True,
        "active": state["active"],
        "nb_lus": state["nb_lus"],
        "pending_lines": len(state["pending"]),
        "last_error": state["last_error"],
    }


@router.post("/{concours_id}/benzing/stop-watch")
async def benzing_stop(concours_id: str):
    watcher = _benzing_watchers.get(concours_id)
    if watcher:
        watcher["state"]["active"] = False
        _benzing_watchers.pop(concours_id, None)
    return {"active": False, "message": "Watch arrêté"}


@router.post("/{concours_id}/benzing/flush", response_model=ImportArriveeResult)
async def benzing_flush(concours_id: str, db: AsyncSession = Depends(get_db)):
    """Traite les lignes Benzing en attente et les applique aux engagements.
    Format Benzing attendu : matricule;date;heure;code_pigeon
    """
    watcher = _benzing_watchers.get(concours_id)
    if not watcher:
        return ImportArriveeResult(traites=0, non_trouves=[], erreurs=["Watch non actif"])

    rc = await db.execute(select(Concours).where(Concours.id == concours_id))
    concours = rc.scalar_one_or_none()
    if not concours:
        raise HTTPException(status_code=404, detail="Concours non trouvé")

    eng_result = await db.execute(
        select(ConcoursEngagement)
        .options(selectinload(ConcoursEngagement.pigeon))
        .where(ConcoursEngagement.concours_id == concours_id)
    )
    engs_by_matricule = {
        e.pigeon.matricule: e for e in eng_result.scalars().all() if e.pigeon
    }

    state = watcher["state"]
    pending = state["pending"][:]
    state["pending"].clear()

    traites, non_trouves, erreurs = 0, [], []

    for line in pending:
        parts = line.split(";")
        if len(parts) < 3:
            erreurs.append(f"Ligne Benzing invalide : {line}")
            continue
        matricule = parts[0].strip()
        heure = parts[2].strip()
        correction = 0

        eng = engs_by_matricule.get(matricule)
        if not eng:
            non_trouves.append(matricule)
            continue

        vitesse = _calculer_vitesse(concours, heure, correction)
        eng.heure_arrivee = heure
        eng.correction_horloge_sec = correction
        eng.vitesse_m_min = vitesse
        eng.source_arrivee = "benzing"

        if concours.session_id:
            tr = await _upsert_training_result(
                db, concours.session_id, eng.pigeon_id,
                return_time=(round(concours.distance_m / vitesse, 2) if concours.distance_m and vitesse else None),
                internal_rank=eng.classement_officiel,
                existing_result_id=eng.training_result_id,
            )
            eng.training_result_id = tr.id

        state["nb_lus"] += 1
        traites += 1

    await db.commit()
    return ImportArriveeResult(traites=traites, non_trouves=non_trouves, erreurs=erreurs)


# ═══════════════════════════════════════════════════════════════════════════════
# FEUILLE D'ENGAGEMENT (JSON pour impression frontend)
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/{concours_id}/feuille-engagement")
async def feuille_engagement(concours_id: str, db: AsyncSession = Depends(get_db)):
    r = await db.execute(
        select(Concours)
        .options(selectinload(Concours.engagements).selectinload(ConcoursEngagement.pigeon))
        .where(Concours.id == concours_id)
    )
    concours = r.scalar_one_or_none()
    if not concours:
        raise HTTPException(status_code=404, detail="Concours non trouvé")

    # Regrouper par catégorie
    from collections import defaultdict
    by_cat: dict = defaultdict(list)
    for eng in concours.engagements:
        p = eng.pigeon
        by_cat[eng.categorie.value if eng.categorie else ""].append({
            "matricule": p.matricule if p else "",
            "annee": p.annee_naissance if p else "",
            "couleur": p.couleur_plumage if p else "",
            "case": p.colombier_case if p else "",
            "categorie": eng.categorie.value if eng.categorie else "",
            "mise": eng.mise,
            "eligible": eng.eligible,
            "alertes": eng.raison_ineligibilite or "",
        })

    return {
        "concours": {
            "nom": concours.nom,
            "date": str(concours.date),
            "lieu_lacher": concours.lieu_lacher,
            "distance_km": round(concours.distance_m / 1000, 1) if concours.distance_m else None,
            "heure_lacher": concours.heure_lacher,
        },
        "nb_total": len(concours.engagements),
        "par_categorie": dict(by_cat),
    }
